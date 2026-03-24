"""Generic CSV and Excel export service for tabular data."""

import csv
from io import BytesIO, StringIO

import openpyxl


class TabularExportService:
    """Generates CSV and multi-sheet Excel files from tabular data."""

    @staticmethod
    def to_csv(headers: list[str], rows: list[list]) -> BytesIO:
        """Export a single table as CSV.

        Args:
            headers: Column header names.
            rows: List of row data (each row is a list of values).

        Returns:
            BytesIO buffer containing UTF-8 encoded CSV with BOM for Excel compatibility.
        """
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)
        # Add UTF-8 BOM so Excel opens the file with correct encoding
        buffer = BytesIO(b"\xef\xbb\xbf" + output.getvalue().encode("utf-8"))
        buffer.seek(0)
        return buffer

    @staticmethod
    def to_excel(sheets: dict[str, tuple[list[str], list[list]]]) -> BytesIO:
        """Export multiple tables as a multi-sheet XLSX file.

        Args:
            sheets: Dict mapping sheet name to (headers, rows) tuple.
                    Sheet names are truncated to 31 chars (Excel limit).

        Returns:
            BytesIO buffer containing the XLSX file.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default empty sheet

        for name, (headers, rows) in sheets.items():
            ws = wb.create_sheet(title=name[:31])
            ws.append(headers)
            for row in rows:
                ws.append(row)

            # Auto-width columns (approximate)
            for col_idx, header in enumerate(headers, 1):
                max_len = len(str(header))
                for row in rows[:50]:  # sample first 50 rows
                    if col_idx <= len(row):
                        max_len = max(max_len, len(str(row[col_idx - 1])))
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 40)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
