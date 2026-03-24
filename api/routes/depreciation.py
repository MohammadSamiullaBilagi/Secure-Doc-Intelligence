"""Depreciation Calculator endpoints — upload fixed asset register, compute IT Act + Companies Act depreciation."""

import asyncio
import logging
import re
from datetime import datetime, date
from pathlib import Path
from uuid import uuid4, UUID
from typing import Annotated
from io import BytesIO

from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc

from api.dependencies import get_current_user, require_starter
from api.rate_limit import limiter
from db.database import get_db
from db.models.core import User, DepreciationAnalysis
from db.models.billing import CreditActionType
from services.credits_service import CreditsService
from services.storage import get_storage
from services.tabular_export_service import TabularExportService

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/depreciation", tags=["depreciation"])

BASE_SESSIONS_DIR = Path("user_sessions")

_FY_PATTERN = re.compile(r"^(\d{4})-(\d{2})$")

_ALLOWED_EXTENSIONS = {".pdf", ".xlsx", ".xls"}


def _dep_dir(user_id: str) -> Path:
    d = BASE_SESSIONS_DIR / user_id / "depreciation"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _validate_fy(fy: str) -> str:
    """Validate FY format like '2025-26'. Returns the FY string or raises."""
    m = _FY_PATTERN.match(fy)
    if not m:
        raise HTTPException(status_code=400, detail="FY must be in format YYYY-YY (e.g. 2025-26)")
    start_year = int(m.group(1))
    end_part = int(m.group(2))
    expected_end = (start_year + 1) % 100
    if end_part != expected_end:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid FY: {fy}. Second part should be {expected_end:02d}",
        )
    return fy


def _extract_text_from_excel(file_path: str) -> str:
    """Extract text from Excel file as pipe-delimited rows."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            lines.append(" | ".join(cells))
    wb.close()
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/compute")
@limiter.limit("5/minute")
async def compute_depreciation(
    request: Request,
    file: UploadFile = File(..., description="Fixed asset register PDF or Excel"),
    fy: str = Form(..., description="Financial year e.g. 2025-26"),
    tax_rate: float = Form(0.25, description="Corporate tax rate for deferred tax (0.0-1.0)"),
    client_id: str = Form(None, description="Optional client UUID"),
    current_user: Annotated[User, Depends(require_starter)] = None,
    db: AsyncSession = Depends(get_db),
):
    """Upload a fixed asset register (PDF/Excel) and compute depreciation schedules."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    fy = _validate_fy(fy)

    # Validate file extension
    ext = Path(file.filename).suffix.lower()
    if ext not in _ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Only PDF and Excel (.xlsx, .xls) files are supported")

    # Validate tax rate
    if not (0.0 <= tax_rate <= 1.0):
        raise HTTPException(status_code=400, detail="Tax rate must be between 0.0 and 1.0")

    # Deduct credits
    await CreditsService.check_and_deduct(
        current_user.id,
        CreditActionType.DEPRECIATION_CALC,
        db,
        description=f"Depreciation calculation: FY {fy}",
    )

    # Save file via storage adapter
    storage = get_storage()
    analysis_id = uuid4()
    saved_filename = f"{analysis_id}_{file.filename}"
    storage_key = f"{current_user.id}/depreciation/{saved_filename}"

    content = await file.read()
    storage.save(storage_key, content)

    # Extract text (needs local path)
    file_path = storage.local_path(storage_key)
    try:
        if ext == ".pdf":
            import pymupdf
            doc = pymupdf.open(str(file_path))
            raw_text = ""
            for page in doc:
                raw_text += page.get_text()
            doc.close()
        else:
            raw_text = await asyncio.to_thread(_extract_text_from_excel, str(file_path))
    except Exception as e:
        logger.error(f"Text extraction failed: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to extract text from file: {e}")

    if not raw_text.strip():
        raise HTTPException(
            status_code=400,
            detail="Could not extract any text from the file. It may be scanned/image-based.",
        )

    # Extract assets via LLM
    from services.depreciation_service import DepreciationService
    svc = DepreciationService()

    try:
        assets = await asyncio.to_thread(svc.extract_asset_register, raw_text)
    except Exception as e:
        logger.error(f"Asset extraction failed: {e}")
        raise HTTPException(status_code=500, detail=f"Asset extraction failed: {e}")

    if not assets:
        raise HTTPException(
            status_code=400,
            detail="Could not extract any fixed assets from the file. Check file format.",
        )

    # Parse FY dates
    start_year = int(fy[:4])
    fy_start = date(start_year, 4, 1)
    fy_end = date(start_year + 1, 3, 31)

    # Compute depreciation schedules
    it_result = await asyncio.to_thread(svc.compute_it_act_depreciation, assets, fy_start, fy_end)
    ca_result = await asyncio.to_thread(svc.compute_companies_act_depreciation, assets, fy_start, fy_end)
    dt_result = svc.compute_deferred_tax(it_result, ca_result, tax_rate)

    total_cost = sum(a.get("cost", 0.0) for a in assets)

    result_data = {
        "it_act": it_result,
        "companies_act": ca_result,
        "deferred_tax": dt_result,
        "assets_extracted": len(assets),
        "total_cost": round(total_cost, 2),
    }

    # Save to DB
    analysis = DepreciationAnalysis(
        id=analysis_id,
        user_id=current_user.id,
        client_id=UUID(client_id) if client_id else None,
        filename=saved_filename,
        fy=fy,
        tax_rate=tax_rate,
        status="completed",
        result_json=result_data,
        total_assets=len(assets),
        total_cost=round(total_cost, 2),
        it_act_depreciation=it_result["total_depreciation"],
        ca_depreciation=ca_result["total_depreciation"],
        timing_difference=dt_result["timing_difference"],
        deferred_tax_amount=dt_result["deferred_tax_amount"],
    )
    db.add(analysis)
    await db.commit()

    return {
        "analysis_id": str(analysis_id),
        "status": "completed",
        "fy": fy,
        "total_assets": len(assets),
        "total_cost": round(total_cost, 2),
        "it_act_depreciation": it_result["total_depreciation"],
        "ca_depreciation": ca_result["total_depreciation"],
        "timing_difference": dt_result["timing_difference"],
        "deferred_tax_amount": dt_result["deferred_tax_amount"],
        "deferred_tax_type": dt_result["deferred_tax_type"],
    }


@router.get("/history")
async def get_depreciation_history(
    current_user: Annotated[User, Depends(require_starter)] = None,
    db: AsyncSession = Depends(get_db),
    limit: int = 20,
    offset: int = 0,
    client_id: str = Query(None, description="Filter by client UUID"),
):
    """Return past depreciation analyses for the user (paginated)."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    query = (
        select(DepreciationAnalysis)
        .where(DepreciationAnalysis.user_id == current_user.id)
    )
    if client_id:
        query = query.where(DepreciationAnalysis.client_id == UUID(client_id))
    query = query.order_by(desc(DepreciationAnalysis.created_at)).limit(limit).offset(offset)

    result = await db.execute(query)
    analyses = result.scalars().all()

    return [
        {
            "analysis_id": str(a.id),
            "filename": a.filename,
            "fy": a.fy,
            "status": a.status,
            "client_id": str(a.client_id) if a.client_id else None,
            "total_assets": a.total_assets,
            "total_cost": a.total_cost,
            "it_act_depreciation": a.it_act_depreciation,
            "ca_depreciation": a.ca_depreciation,
            "timing_difference": a.timing_difference,
            "deferred_tax_amount": a.deferred_tax_amount,
            "created_at": a.created_at.isoformat() if a.created_at else None,
        }
        for a in analyses
    ]


@router.get("/{analysis_id}/csv")
async def download_depreciation_csv(
    analysis_id: str,
    sheet: str = Query("IT Act Blocks", description="Table to export: IT Act Blocks, Companies Act Assets, Deferred Tax"),
    current_user: Annotated[User, Depends(require_starter)] = None,
    db: AsyncSession = Depends(get_db),
):
    """Download a depreciation analysis table as CSV."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    result = await db.execute(
        select(DepreciationAnalysis).where(
            DepreciationAnalysis.id == UUID(analysis_id),
            DepreciationAnalysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Depreciation analysis not found")
    if analysis.status != "completed" or not analysis.result_json:
        raise HTTPException(status_code=400, detail="Analysis is not completed yet")

    tables = _extract_dep_tables(analysis.result_json)
    if sheet not in tables:
        raise HTTPException(status_code=400, detail=f"Invalid sheet '{sheet}'. Available: {', '.join(tables.keys())}")

    headers, rows = tables[sheet]
    csv_buffer = TabularExportService.to_csv(headers, rows)
    safe_sheet = sheet.lower().replace(" ", "_")
    filename = f"depreciation_{safe_sheet}_{analysis.fy}_{str(analysis.id)[:8]}.csv"
    return StreamingResponse(
        csv_buffer, media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{analysis_id}/excel")
async def download_depreciation_excel(
    analysis_id: str,
    current_user: Annotated[User, Depends(require_starter)] = None,
    db: AsyncSession = Depends(get_db),
):
    """Download all depreciation tables as multi-sheet Excel."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    result = await db.execute(
        select(DepreciationAnalysis).where(
            DepreciationAnalysis.id == UUID(analysis_id),
            DepreciationAnalysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Depreciation analysis not found")
    if analysis.status != "completed" or not analysis.result_json:
        raise HTTPException(status_code=400, detail="Analysis is not completed yet")

    tables = _extract_dep_tables(analysis.result_json)
    excel_buffer = TabularExportService.to_excel(tables)
    filename = f"depreciation_{analysis.fy}_{str(analysis.id)[:8]}.xlsx"
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{analysis_id}")
async def get_depreciation_analysis(
    analysis_id: str,
    current_user: Annotated[User, Depends(require_starter)] = None,
    db: AsyncSession = Depends(get_db),
):
    """Return full depreciation analysis result."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    result = await db.execute(
        select(DepreciationAnalysis).where(
            DepreciationAnalysis.id == UUID(analysis_id),
            DepreciationAnalysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Depreciation analysis not found")

    return {
        "analysis_id": str(analysis.id),
        "filename": analysis.filename,
        "fy": analysis.fy,
        "status": analysis.status,
        "tax_rate": analysis.tax_rate,
        "total_assets": analysis.total_assets,
        "total_cost": analysis.total_cost,
        "it_act_depreciation": analysis.it_act_depreciation,
        "ca_depreciation": analysis.ca_depreciation,
        "timing_difference": analysis.timing_difference,
        "deferred_tax_amount": analysis.deferred_tax_amount,
        "created_at": analysis.created_at.isoformat() if analysis.created_at else None,
        "result": analysis.result_json,
    }


@router.get("/{analysis_id}/report")
async def generate_depreciation_report(
    analysis_id: str,
    current_user: Annotated[User, Depends(require_starter)] = None,
    db: AsyncSession = Depends(get_db),
):
    """Generate and download a PDF report for the depreciation analysis."""
    if not current_user:
        raise HTTPException(status_code=401, detail="Unauthorized")

    result = await db.execute(
        select(DepreciationAnalysis).where(
            DepreciationAnalysis.id == UUID(analysis_id),
            DepreciationAnalysis.user_id == current_user.id,
        )
    )
    analysis = result.scalar_one_or_none()
    if not analysis:
        raise HTTPException(status_code=404, detail="Depreciation analysis not found")

    if analysis.status != "completed" or not analysis.result_json:
        raise HTTPException(status_code=400, detail="Analysis is not completed yet")

    pdf_buffer = await asyncio.to_thread(_generate_depreciation_pdf, analysis)

    filename = f"depreciation_{analysis.fy}_{str(analysis.id)[:8]}.pdf"
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# PDF Report Generation
# ---------------------------------------------------------------------------

def _generate_depreciation_pdf(analysis: DepreciationAnalysis) -> BytesIO:
    """Build PDF report from depreciation analysis data using fpdf2."""
    from fpdf import FPDF
    from services.report_service import _sanitize_text

    result = analysis.result_json
    it_act = result.get("it_act", {})
    ca_act = result.get("companies_act", {})
    dt = result.get("deferred_tax", {})

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.set_left_margin(15)
    pdf.set_right_margin(15)
    pdf.add_page()

    effective_w = pdf.w - pdf.l_margin - pdf.r_margin

    # Title
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, "Depreciation Schedule Report", ln=True, align="C")
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 12)
    pdf.cell(0, 8, _sanitize_text(f"Financial Year: {analysis.fy}"), ln=True, align="C")
    pdf.set_font("Helvetica", "", 9)
    pdf.cell(0, 6, _sanitize_text(f"File: {analysis.filename}"), ln=True, align="C")
    pdf.cell(0, 6, _sanitize_text(f"Tax Rate: {analysis.tax_rate * 100:.1f}%"), ln=True, align="C")
    pdf.ln(5)

    # Summary
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, "Summary", ln=True)
    pdf.set_font("Helvetica", "", 10)
    stats = [
        f"Total Assets: {result.get('assets_extracted', 0)}",
        f"Total Cost: Rs. {result.get('total_cost', 0):,.2f}",
        f"IT Act Depreciation (WDV): Rs. {it_act.get('total_depreciation', 0):,.2f}",
        f"Companies Act Depreciation (SLM): Rs. {ca_act.get('total_depreciation', 0):,.2f}",
        f"Timing Difference: Rs. {dt.get('timing_difference', 0):,.2f}",
        f"Deferred Tax ({dt.get('deferred_tax_type', 'NIL')}): Rs. {dt.get('deferred_tax_amount', 0):,.2f}",
    ]
    for s in stats:
        pdf.cell(0, 6, _sanitize_text(s), ln=True)
    pdf.ln(5)

    # Section 1: IT Act Depreciation Schedule (block-wise)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, "1. IT Act Depreciation Schedule (WDV Method - Section 32)", ln=True)
    blocks = it_act.get("blocks", {})
    if blocks:
        pdf.set_font("Helvetica", "B", 6.5)
        pdf.set_fill_color(240, 240, 240)
        col_w = [30, 12, 22, 22, 18, 20, 22, 22, 22]
        headers = ["Block", "Rate%", "Opening", "Additions", "Half-Yr", "Disposals", "Net Block", "Dep.", "Closing"]
        for i, h in enumerate(headers):
            pdf.cell(col_w[i], 6, h, border=1, fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", "", 6)
        for bk, blk in blocks.items():
            label = bk.replace("_", " ").title()[:20]
            pdf.cell(col_w[0], 5, _sanitize_text(label), border=1)
            pdf.cell(col_w[1], 5, f"{blk.get('rate', 0) * 100:.0f}%", border=1)
            pdf.cell(col_w[2], 5, f"{blk.get('opening_wdv', 0):,.0f}", border=1)
            pdf.cell(col_w[3], 5, f"{blk.get('additions', 0):,.0f}", border=1)
            pdf.cell(col_w[4], 5, f"{blk.get('half_year_additions', 0):,.0f}", border=1)
            pdf.cell(col_w[5], 5, f"{blk.get('disposals', 0):,.0f}", border=1)
            pdf.cell(col_w[6], 5, f"{blk.get('net_block', 0):,.0f}", border=1)
            pdf.cell(col_w[7], 5, f"{blk.get('depreciation', 0):,.0f}", border=1)
            pdf.cell(col_w[8], 5, f"{blk.get('closing_wdv', 0):,.0f}", border=1)
            pdf.ln()

        # Totals row
        pdf.set_font("Helvetica", "B", 6)
        pdf.cell(col_w[0], 5, "TOTAL", border=1, fill=True)
        pdf.cell(col_w[1], 5, "", border=1, fill=True)
        pdf.cell(col_w[2], 5, f"{sum(b.get('opening_wdv', 0) for b in blocks.values()):,.0f}", border=1, fill=True)
        pdf.cell(col_w[3], 5, f"{sum(b.get('additions', 0) for b in blocks.values()):,.0f}", border=1, fill=True)
        pdf.cell(col_w[4], 5, f"{sum(b.get('half_year_additions', 0) for b in blocks.values()):,.0f}", border=1, fill=True)
        pdf.cell(col_w[5], 5, f"{sum(b.get('disposals', 0) for b in blocks.values()):,.0f}", border=1, fill=True)
        pdf.cell(col_w[6], 5, f"{sum(b.get('net_block', 0) for b in blocks.values()):,.0f}", border=1, fill=True)
        pdf.cell(col_w[7], 5, f"{it_act.get('total_depreciation', 0):,.0f}", border=1, fill=True)
        pdf.cell(col_w[8], 5, f"{sum(b.get('closing_wdv', 0) for b in blocks.values()):,.0f}", border=1, fill=True)
        pdf.ln()
    pdf.ln(5)

    # Section 2: Companies Act Schedule (asset-wise)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, "2. Companies Act 2013 Depreciation (SLM - Schedule II)", ln=True)
    ca_assets = ca_act.get("assets", [])
    if ca_assets:
        pdf.set_font("Helvetica", "B", 6)
        pdf.set_fill_color(240, 240, 240)
        ca_cols = [35, 20, 18, 12, 18, 18, 20, 18]
        ca_headers = ["Asset", "Cost", "Residual", "Life", "Annual", "FY Dep", "Accum.", "NBV"]
        for i, h in enumerate(ca_headers):
            pdf.cell(ca_cols[i], 6, h, border=1, fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", "", 5.5)
        for asset in ca_assets[:100]:
            desc = str(asset.get("description", ""))[:24]
            pdf.cell(ca_cols[0], 5, _sanitize_text(desc), border=1)
            pdf.cell(ca_cols[1], 5, f"{asset.get('cost', 0):,.0f}", border=1)
            pdf.cell(ca_cols[2], 5, f"{asset.get('residual', 0):,.0f}", border=1)
            pdf.cell(ca_cols[3], 5, f"{asset.get('useful_life', 0)}y", border=1)
            pdf.cell(ca_cols[4], 5, f"{asset.get('annual_dep', 0):,.0f}", border=1)
            pdf.cell(ca_cols[5], 5, f"{asset.get('fy_dep', 0):,.0f}", border=1)
            pdf.cell(ca_cols[6], 5, f"{asset.get('accumulated_total', 0):,.0f}", border=1)
            pdf.cell(ca_cols[7], 5, f"{asset.get('nbv', 0):,.0f}", border=1)
            pdf.ln()

        if len(ca_assets) > 100:
            pdf.set_font("Helvetica", "I", 7)
            pdf.cell(0, 5, f"... and {len(ca_assets) - 100} more assets", ln=True)

        # CA totals
        pdf.set_font("Helvetica", "B", 6)
        pdf.cell(ca_cols[0], 5, "TOTAL", border=1, fill=True)
        pdf.cell(ca_cols[1], 5, f"{sum(a.get('cost', 0) for a in ca_assets):,.0f}", border=1, fill=True)
        pdf.cell(ca_cols[2], 5, "", border=1, fill=True)
        pdf.cell(ca_cols[3], 5, "", border=1, fill=True)
        pdf.cell(ca_cols[4], 5, "", border=1, fill=True)
        pdf.cell(ca_cols[5], 5, f"{ca_act.get('total_depreciation', 0):,.0f}", border=1, fill=True)
        pdf.cell(ca_cols[6], 5, "", border=1, fill=True)
        pdf.cell(ca_cols[7], 5, f"{ca_act.get('total_nbv', 0):,.0f}", border=1, fill=True)
        pdf.ln()
    pdf.ln(5)

    # Section 3: Comparison + Timing Difference
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, "3. IT Act vs Companies Act - Block-wise Comparison", ln=True)
    comparisons = dt.get("block_wise_comparison", [])
    if comparisons:
        pdf.set_font("Helvetica", "B", 7)
        pdf.set_fill_color(240, 240, 240)
        cmp_cols = [35, 30, 30, 30, 20]
        cmp_headers = ["Block", "IT Act Dep.", "CA Dep.", "Difference", "Type"]
        for i, h in enumerate(cmp_headers):
            pdf.cell(cmp_cols[i], 6, h, border=1, fill=True)
        pdf.ln()

        pdf.set_font("Helvetica", "", 7)
        for c in comparisons:
            label = c.get("block", "").replace("_", " ").title()[:24]
            diff = c.get("difference", 0)
            pdf.cell(cmp_cols[0], 5, _sanitize_text(label), border=1)
            pdf.cell(cmp_cols[1], 5, f"{c.get('it_act_depreciation', 0):,.2f}", border=1)
            pdf.cell(cmp_cols[2], 5, f"{c.get('ca_depreciation', 0):,.2f}", border=1)
            if diff > 0:
                pdf.set_text_color(200, 0, 0)
            elif diff < 0:
                pdf.set_text_color(0, 100, 0)
            pdf.cell(cmp_cols[3], 5, f"{diff:,.2f}", border=1)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(cmp_cols[4], 5, c.get("type", "NIL"), border=1)
            pdf.ln()
    pdf.ln(5)

    # Section 4: Deferred Tax Computation
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 9, "4. Deferred Tax Computation", ln=True)
    pdf.set_font("Helvetica", "", 10)
    dt_lines = [
        f"IT Act Total Depreciation: Rs. {dt.get('it_act_depreciation', 0):,.2f}",
        f"Companies Act Total Depreciation: Rs. {dt.get('ca_depreciation', 0):,.2f}",
        f"Timing Difference: Rs. {dt.get('timing_difference', 0):,.2f}",
        f"Applicable Tax Rate: {dt.get('tax_rate', 0.25) * 100:.1f}%",
        f"Deferred Tax Amount: Rs. {dt.get('deferred_tax_amount', 0):,.2f}",
        f"Type: {dt.get('deferred_tax_type', 'NIL')} ({'Deferred Tax Liability' if dt.get('deferred_tax_type') == 'DTL' else 'Deferred Tax Asset' if dt.get('deferred_tax_type') == 'DTA' else 'No Difference'})",
    ]
    for line in dt_lines:
        pdf.cell(0, 7, _sanitize_text(line), ln=True)
    pdf.ln(5)

    # Section 5: Capital Gains on Block Disposals
    capital_gain_blocks = {bk: blk for bk, blk in blocks.items() if blk.get("capital_gain", 0) > 0}
    if capital_gain_blocks:
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 9, "5. Capital Gains on Block Disposals (STCG)", ln=True)
        pdf.set_font("Helvetica", "", 10)
        for bk, blk in capital_gain_blocks.items():
            label = bk.replace("_", " ").title()
            pdf.cell(0, 7, _sanitize_text(
                f"{label}: Disposal exceeded block value — STCG Rs. {blk['capital_gain']:,.2f}"
            ), ln=True)
        pdf.ln(5)

    # Disclaimer + footer
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 8, "Disclaimer", ln=True)
    pdf.set_font("Helvetica", "", 7)
    disclaimers = [
        "This report is generated based on the data extracted from the uploaded fixed asset register.",
        "IT Act depreciation uses WDV block method per Section 32. Companies Act uses SLM per Schedule II.",
        "Deferred tax computation is for AS-22/Ind AS 12 disclosure purposes only.",
        "Verify all computations with your Chartered Accountant before finalizing financial statements.",
        "This is not professional advice. Consult a qualified CA for compliance.",
    ]
    for d in disclaimers:
        pdf.cell(0, 4, _sanitize_text(d), ln=True)
    pdf.ln(3)

    pdf.set_font("Helvetica", "I", 8)
    pdf.cell(0, 10, _sanitize_text(
        f"Generated by Secure Doc-Intelligence | {datetime.utcnow().strftime('%Y-%m-%d')}"
    ), ln=True, align="C")

    buffer = BytesIO()
    buffer.write(pdf.output())
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Table Extraction (for CSV/Excel export)
# ---------------------------------------------------------------------------

def _extract_dep_tables(result_json: dict) -> dict:
    """Extract all tables from depreciation result_json for CSV/Excel export."""
    tables = {}

    # IT Act blocks
    it_act = result_json.get("it_act", {})
    blocks = it_act.get("blocks", {})
    headers_it = ["Block", "Rate %", "Opening WDV", "Additions", "Half-Yr Additions",
                  "Disposals", "Net Block", "Depreciation", "Closing WDV"]
    rows_it = []
    for bk, blk in blocks.items():
        rows_it.append([
            bk.replace("_", " ").title(), f"{blk.get('rate', 0) * 100:.0f}%",
            blk.get("opening_wdv", 0), blk.get("additions", 0), blk.get("half_year_additions", 0),
            blk.get("disposals", 0), blk.get("net_block", 0), blk.get("depreciation", 0),
            blk.get("closing_wdv", 0),
        ])
    tables["IT Act Blocks"] = (headers_it, rows_it)

    # Companies Act assets
    ca = result_json.get("companies_act", {})
    ca_assets = ca.get("assets", [])
    headers_ca = ["Asset", "Cost", "Residual Value", "Useful Life", "Annual Depreciation",
                  "FY Depreciation", "Accumulated", "NBV"]
    rows_ca = [[
        a.get("description", ""), a.get("cost", 0), a.get("residual", 0),
        a.get("useful_life", 0), a.get("annual_dep", 0), a.get("fy_dep", 0),
        a.get("accumulated_total", 0), a.get("nbv", 0),
    ] for a in ca_assets]
    tables["Companies Act Assets"] = (headers_ca, rows_ca)

    # Deferred tax
    dt = result_json.get("deferred_tax", {})
    headers_dt = ["Item", "Value"]
    rows_dt = [
        ["IT Act Depreciation", dt.get("it_act_depreciation", 0)],
        ["Companies Act Depreciation", dt.get("ca_depreciation", 0)],
        ["Timing Difference", dt.get("timing_difference", 0)],
        ["Tax Rate", f"{dt.get('tax_rate', 0.25) * 100:.1f}%"],
        ["Deferred Tax Amount", dt.get("deferred_tax_amount", 0)],
        ["Type", dt.get("deferred_tax_type", "NIL")],
    ]
    tables["Deferred Tax"] = (headers_dt, rows_dt)

    return tables
